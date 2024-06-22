import os
import shutil
from time import sleep
import openpyxl as xl
import pyautogui
import pygetwindow as gw

class Config:
    def __init__(self, file_name):
        self.file_name = file_name
        self.data = {}
        self.load()

    def load(self):
        try:
            with open(self.file_name, 'r') as file:
                self.parse(file)
        except FileNotFoundError:
            raise Exception('There is no ' + self.file_name)

    def parse(self, file):
        lines = file.readlines()
        try:
            for line in lines:
                text = line.strip()
                text = text.replace(' ', '')
                text = text.split(':')
                self.data[text[0]] = text[1]
        except IndexError:
            raise Exception(self.file_name + ' Format Error use Key:Value')

    def get_value(self, key):
        return self.data.get(key)


def create_directory(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        raise Exception('Failed to create the directory')


class Xml:
    def __init__(self, file_name):
        self.file_name = file_name + '.xlsx'
        self.wb = None
        self.load()

    def load(self):
        try:
            self.wb = xl.load_workbook(self.file_name)
        except FileNotFoundError:
            raise Exception('There is no ' + self.file_name)


def active_window(process_name):
    if process_name is None:
        raise Exception('ProcessName is None')

    windows = gw.getWindowsWithTitle(process_name)
    if 1 != len(windows):
        raise Exception(process_name + ' is not running')

    window = windows[0]
    if not window.isMaximized:
        window.maximize()

    if not window.isActive:
        window.activate()


def run(save_path, key_word_list):
    active_window(config.get_value('ProcessName'))

    create_directory(save_path)

    shutil.rmtree(save_path)

    for keyWorld in key_word_list:
        create_directory(save_path + keyWorld)

    for sheet_name in xml.wb.sheetnames:
        sheet = xml.wb[sheet_name]

        index = 0
        cell_count = 0
        for row in sheet.iter_rows(min_row=1):
            for cell in row:
                cell_count += 1

        for row in sheet.iter_rows(min_row=1):
            for cell in row:
                index += 1
                print(str(index) + '/' + str(cell_count) + ' ' + cell.value)

                for keyWorld in key_word_list:
                    pyautogui.hotkey('ctrl', 'shift', 'f')
                    pyautogui.press('delete')
                    pyautogui.write(keyWorld + ' ' + cell.value)
                    pyautogui.press('enter')
                    sleep(2)
                    img = pyautogui.screenshot()
                    img.save(save_path + keyWorld + '/' + cell.value + '.png')

try:
    config = Config('Config.txt')
    xml = Xml(config.get_value('ExcelName'))

    key_word_list = config.get_value('KeyWordList')
    if key_word_list is None:
        raise Exception('There is no KeyWordList')
    key_word_list = key_word_list.split(',')

    run(save_path=r'Img/', key_word_list=key_word_list)
except Exception as e:
    print('Error:', e)
    exit(0)

os.system("pause")
